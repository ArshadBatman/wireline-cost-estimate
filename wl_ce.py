import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

st.title("SMARTLog: Wireline Cost Estimator")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

# --- Well A definition (reference well) ---
reference_wells = {
    "Well A": {
        "Package": "Package A",
        "Service": "Standard Wells",
        "Hole Sections": {
            '12.25"': {"Quantity": 2, "Total Months": 1, "Depth": 5500},
            '8.5"': {"Quantity": 2, "Total Months": 1, "Depth": 8000}
        },
        # Map to the special-case group keys you already use in special_cases_map
        "Tool Groups": [
            "PEX-Rt Scanner (150DegC Max)",
            "ECS-NMR (150DegC Max)",
            "Dual-OBMI DSI (150DegC Max)",   # map to DOBMI group in your special_cases_map
            "MDT: LFA-QS-XLD-MIFA-Saturn-2MS (150DegC Max)",
            "XL Rock (150DegC Max)",
            "Pipe Conveyed Logging",
            "FPIT & Back-off services / Drilling ontingent Support Services",
            "Unit, Cables & Conveyance",
            "Personnel",
        ]
    }
}

# --- Reference Well Selector (sidebar) ---
st.sidebar.header("Reference Well Selection")
selected_well = st.sidebar.selectbox("Reference Well", ["None"] + list(reference_wells.keys()), index=0)

if uploaded_file:
    # --- Reset unique tracker when a new file is uploaded ---
    if "last_uploaded_name" not in st.session_state or st.session_state["last_uploaded_name"] != uploaded_file.name:
        st.session_state["unique_tracker"] = set()
        st.session_state["last_uploaded_name"] = uploaded_file.name

    # Optional manual reset button
    if st.sidebar.button("Reset unique-tool usage (AU14, etc.)", key="reset_unique"):
        st.session_state["unique_tracker"] = set()
        st.sidebar.success("Unique-tool tracker cleared.")

    # Read data
    df = pd.read_excel(uploaded_file, sheet_name="Data")

    # Unique tools across sections
    unique_tools = {"AU14: AUX_SURELOC"}

    # --- Well selection (top of sidebar) ---
    st.sidebar.header("Well Selection")
    # keep original selectbox too, but show chosen reference well state
    well_option = selected_well  # "None" or "Well A"

    # --- Dynamic Hole Section Setup ---
    st.sidebar.header("Hole Sections Setup")
    # If Well A selected, default number of sections = 2, else default 2 (same)
    default_num_sections = 2 if well_option == "Well A" else 2
    num_sections = st.sidebar.number_input("Number of Hole Sections", min_value=1, max_value=5, value=default_num_sections, step=1)

    # If Well A selected, use its hole sizes as defaults; otherwise use calculated defaults
    hole_sizes = []
    if well_option == "Well A":
        # Use exact keys from reference_wells (keeps the inch formatting)
        hole_sizes = list(reference_wells["Well A"]["Hole Sections"].keys())
        # If user changed num_sections and it doesn't match default, allow editing below; but default hole sizes are used initially
        # If UI number_input changed to different value, still show fields for each index—we'll fallback to provided default formatting
        if len(hole_sizes) < num_sections:
            # append generic sizes if necessary
            for i in range(len(hole_sizes), num_sections):
                hole_sizes.append(f'{12.25 - i*3.75:.2f}')
        elif len(hole_sizes) > num_sections:
            hole_sizes = hole_sizes[:num_sections]
    else:
        for i in range(num_sections):
            hole_size = st.sidebar.text_input(f"Hole Section {i+1} Size (inches)", value=f"{12.25 - i*3.75:.2f}")
            hole_sizes.append(hole_size)

    # If Well A selected, prefill session_state defaults for qty/months/depth & package/service
    if well_option == "Well A":
        well_info = reference_wells["Well A"]
        for hole_size in hole_sizes:
            # Use the exact key like '12.25"' or '8.5"'
            if hole_size in well_info["Hole Sections"]:
                info = well_info["Hole Sections"][hole_size]
                # Set session state defaults (will be used by number_input default values below)
                st.session_state.setdefault(f"qty_{hole_size}", info.get("Quantity", 2))
                st.session_state.setdefault(f"months_{hole_size}", info.get("Total Months", 1))
                st.session_state.setdefault(f"depth_{hole_size}", info.get("Depth", 5500))
                # set package/service default placeholders (we'll apply when options exist)
                st.session_state.setdefault(f"pkg_default_for_{hole_size}", well_info["Package"])
                st.session_state.setdefault(f"svc_default_for_{hole_size}", well_info["Service"])
    # Create dynamic tabs
    tabs = st.tabs([f'{hs}" Hole Section' for hs in hole_sizes])
    section_totals = {}
    all_calc_dfs_for_excel = []  # store data for Excel download

    # --- Loop for each hole section ---
    for tab, hole_size in zip(tabs, hole_sizes):
        with tab:
            st.header(f'{hole_size}" Hole Section')

            # Sidebar inputs per section
            st.sidebar.subheader(f"Inputs for {hole_size}\" Section")

            # --- Force default quantity = 0 for Well A 8.5" Hole Section ---
            if well_option == "Well A" and hole_size == '8.5"':
                st.session_state[f"qty_{hole_size}"] = 0  # overwrite session state
                default_qty = 0
            else:
                default_qty = st.session_state.get(f"qty_{hole_size}", 2)
            
            quantity_tools = st.sidebar.number_input(
                f"Quantity of Tools ({hole_size})",
                min_value=0,
                value=default_qty,
                key=f"qty_{hole_size}"
            )

            total_days = st.sidebar.number_input(
                f"Total Days ({hole_size})", min_value=0, value=st.session_state.get(f"days_{hole_size}", 0), key=f"days_{hole_size}"
            )
            total_months = st.sidebar.number_input(
                f"Total Months ({hole_size})",
                min_value=0,
                value=st.session_state.get(f"months_{hole_size}", 1),
                key=f"months_{hole_size}"
            )
            total_depth = st.sidebar.number_input(
                f"Total Depth (ft) ({hole_size})",
                min_value=0,
                value=st.session_state.get(f"depth_{hole_size}", 5500),
                key=f"depth_{hole_size}"
            )
            total_survey = st.sidebar.number_input(
                f"Total Survey (ft) ({hole_size})", min_value=0, value=st.session_state.get(f"survey_{hole_size}", 0), key=f"survey_{hole_size}"
            )
            total_hours = st.sidebar.number_input(
                f"Total Hours ({hole_size})", min_value=0, value=st.session_state.get(f"hours_{hole_size}", 0), key=f"hours_{hole_size}"
            )
            # disc stored in session state earlier may be fraction or percent — ensure consistent default
            disc_default = st.session_state.get(f"disc_{hole_size}", 0.0)
            if disc_default > 1.0:
                # if earlier stored as percent e.g. 5.0, convert to fraction
                disc_default_fraction = disc_default / 100.0
            else:
                disc_default_fraction = disc_default
            discount = st.sidebar.number_input(
                f"Discount (%) ({hole_size})",
                min_value=0.0,
                max_value=100.0,
                value=disc_default_fraction * 100,
                key=f"disc_{hole_size}"
            ) / 100.0

           # --- Package & Service ---
            st.subheader("Select Package")
            package_options = df["Package"].dropna().unique().tolist()
            
            # --- Default package logic for Well A ---
            default_pkg = None
            if well_option == "Well A":
                pkg_candidate = reference_wells["Well A"]["Package"]
                if pkg_candidate in package_options:
                    default_pkg = pkg_candidate
            
            if default_pkg:
                try:
                    selected_package = st.selectbox(
                        "Choose Package", package_options, index=package_options.index(default_pkg), key=f"pkg_{hole_size}"
                    )
                except Exception:
                    selected_package = st.selectbox("Choose Package", package_options, key=f"pkg_{hole_size}")
            else:
                selected_package = st.selectbox("Choose Package", package_options, key=f"pkg_{hole_size}")
            
            package_df = df[df["Package"] == selected_package]
            
            # --- Service Name selection ---
            st.subheader("Select Service Name")
            service_options = package_df["Service Name"].dropna().unique().tolist()
            service_options = [svc for svc in service_options if svc.strip() != ""]  # remove empty strings
            
           
            # Always allow blank option
            if "" not in service_options:
                service_options.append("")
            
            # Default service if Well A
            default_svc = None
            if well_option == "Well A":
                svc_candidate = reference_wells["Well A"]["Service"]  # usually "Standard Wells"
                if svc_candidate in service_options:
                    default_svc = svc_candidate
            
            # Select Service Name
            if default_svc:
                try:
                    selected_service = st.selectbox(
                        "Choose Service Name", service_options, index=service_options.index(default_svc), key=f"svc_{hole_size}"
                    )
                except Exception:
                    selected_service = st.selectbox("Choose Service Name", service_options, key=f"svc_{hole_size}")
            else:
                # fallback to normal selectbox
                if len(service_options) == 0:
                    selected_service = st.selectbox("Choose Service Name", [""], index=0, key=f"svc_{hole_size}")
                else:
                    selected_service = st.selectbox("Choose Service Name", service_options, key=f"svc_{hole_size}")
            
            # --- Filter DataFrame based on selected service (including blanks) ---
            df_service = package_df[
                (package_df["Service Name"] == selected_service) | 
                (package_df["Service Name"].isna()) | 
                (package_df["Service Name"] == "")
            ]



            # --- Tool selection with special cases ---
            # Build the special_cases_map exactly as in your original code so the groups match
            code_list = df_service["Specification 1"].dropna().unique().tolist()
            special_cases_map = {
                "STANDARD WELLS": {
                    "PEX-Rt Scanner (150DegC Max)": ["AU14: AUX_SURELOC","NE1: NEUT_THER","DE1: DENS_FULL","RE4: RES_ANIS"],
                    "PEX-AIT (150DegC Max)": ["AU14: AUX_SURELOC","GR1: GR_TOTL","NE1: NEUT_THER","DE1: DENS_FULL","RE1: RES_INDU"],
                    "PEX-AIT-DSI (150DegC Max)": ["AU14: AUX_SURELOC","GR1: GR_TOTL","NE1: NEUT_THER","DE1: DENS_FULL","RE1: RES_INDU",
                                                 "AU3:AUX_INCL", "AU2: AUX_PCAL", "AU2: AUX_PCAL", "AC3: ACOU_3", "PP7: PROC_PETR7", "PA7: PROC_ACOU6",
                                                 "PA11: PROC_ACOU13", "PA12: PROC_ACOU14"],
                    "Dual-OBMI DSI (150DegC Max)": ["AU14: AUX_SURELOC","GR1: GR_TOTL","AU3: AUX_INCL","AC3: ACOU_3",
                                            "AU2: AUX_PCAL","AU2: AUX_PCAL","PP7: PROC_PETR7","PA7: PROC_ACOU6",
                                            "PA11: PROC_ACOU13","PA12: PROC_ACOU14","IM3: IMAG_SOBM","PI1: PROC_IMAG1",
                                            "PI2: PROC_IMAG2","PI7: PROC_IMAG7"],
                    "Dual OBMI-Sonic Scanner (150DegC Max) ": ["AU14: AUX_SURELOC","GR1: GR_TOTL","AU3: AUX_INCL","AC4: ACOU_ADD1",
                                            "AU2: AUX_PCAL","AU2: AUX_PCAL","PP7: PROC_PETR7","PA7: PROC_ACOU6",
                                            "PA11: PROC_ACOU13","PA12: PROC_ACOU14","IM3: IMAG_SOBM","PI1: PROC_IMAG1",
                                            "PI2: PROC_IMAG2","PI7: PROC_IMAG7","PI8: PROC_IMAG8","PI9: PROC_IMAG9",
                                            "PI12: PROC_IMAG12","PI13: PROC_IMAG13"],
                    "MDT: LFA-QS-XLD-MIFA-Saturn-2MS (150DegC Max)": ["AU14: AUX_SURELOC","FP25: FPS_SCAR","FP25: FPS_SCAR",
                                                                      "FP18: FPS_SAMP","FP19: FPS_SPHA","FP23: FPS_TRA",
                                                                      "FP24: FPS_TRK","FP28: FPS_FCHA_1","FP33: FPS_FCHA_6",
                                                                      "FP34: FPS_FCHA_7","FP14: FPS_PUMP","FP14: FPS_PUMP",
                                                                      "FP42: FPS_PROB_XLD","FP11: FPS_PROB_FO","FP26: FPS_FCON",
                                                                      "DT3:RTDT_PER","PPT12: PROC_PT12","FP7: FPS_SPPT_2"],
                    "ECS-NMR (150DegC Max)": ["AU14: AUX_SURELOC","GR1: GR_TOTL","EC1: ES_1","NM1: NMR_1","PN1: PROC_NMR1","PN2: PROC_NMR2","PN6: PROC_NMR6","PE1: PROC_ES1","PP1: PROC_PETR1","PP6: PROC_PETR6", "PN3: PROC_NMR3"],
                    "IBC (PowerFlex)-CBL (150DegC Max)": ["CE1:CES_CBL","CE4:CES_CBI_3","CE6:CES_CBI_5", "DT3:RTDT_PER", "PPT13:PROC_PT13", "DT12:USI-DIG-LP-CET3"],
                    "DSI-QuantaGeo-Rt Scanner (150DegC Max)": ["AU14: AUX_SURELOC","GR1: GR_TOTL","AU3: AUX_INCL", "AC4: ACOU_ADD1", "AC3: ACOU_3", "AU2:AUX_PCAL",
                                                              "AU2:AUX_PCAL", "IM4:IMAG_ADD1","PI1:PROC_IMAG1", "DT4:SONIC-WELL-P/S-DIG", "PI2: PROC_IMAG2", "PI7: PROC_IMAG7",
                                                              "PI8:PROC_IMAG8", "PI9:PROC_IMAG9", "PI12: PROC_IMAG12", "PI13: PROC_IMAG13", "RE4: RES_ANIS"],
                    "Pipe Conveyed Logging": ["CO1: CONV_PCL"],
                    "FPIT & Back-off services / Drilling ontingent Support Services": ["AU7: AUX_SBOX","PC5: PC_10KH2S","PR1: PR_FP","PR2: PR_BO","PR3: PR_TP","AU11: AUX_GRCCL",
                                                                                      "PR7: PR_CST","MS1: MS_PL","MS3: MS_JB"],
                    "Unit, Cables & Conveyance": ["LU1: LUDR_ZON2","CA9: CABL_HSOH_1","CA3: CABL_HSOH","CA8: CABL_STCH_2","DT2:RTDT_SAT"],
                    "XL Rock (150DegC Max)": ["AU14: AUX_SURELOC","SC2: SC_ADD1","SC2: SC_ADD2"],
                    "XL Rock (150DegC Max) With Core Detection": ["AU14: AUX_SURELOC","SC2: SC_ADD1","SC2: SC_ADD2", "SC4: SC_ADD4"],
                    "Personnel": ["PER1:PWFE","PER2:PWSO","PER3:PWOP", "PER4:PWSE"]
                },
                "HT WELLS": {}
            }

            special_cases = special_cases_map.get(selected_service, {})
            code_list_with_special = list(special_cases.keys()) + code_list

            # If Well A is selected, automatically preselect the Well A tool groups
            default_selected_groups = []
            if well_option == "Well A":
                default_groups = reference_wells["Well A"]["Tool Groups"]
                default_selected_groups = [g for g in default_groups if g in code_list_with_special]
            
                # Exclude certain groups for 12.25" Hole Section
                if hole_size == '12.25"':
                    exclude_services_12_25 = [
                        "Pipe Conveyed Logging",
                        "FPIT & Back-off services / Drilling ontingent Support Services",
                        "Unit, Cables & Conveyance",
                        "Personnel"
                    ]
                    default_selected_groups = [g for g in default_selected_groups if g not in exclude_services_12_25]

            selected_codes = st.multiselect("Select Tools (by Specification 1)", code_list_with_special, default=default_selected_groups, key=f"tools_{hole_size}")

            # --- Expand selected special cases ---
            expanded_codes = []
            used_special_cases = []
            for code in selected_codes:
                if code in special_cases:
                    expanded_codes.extend(special_cases[code])
                    used_special_cases.append(code)
                else:
                    expanded_codes.append(code)

            # --- If Well A selected AND special groups were auto-selected above, ensure the mapped codes from special_cases are included even if df_service doesn't contain all codes.
            # df_tools picks only those present in df_service, so Excel/calculation will use what's present.
            df_tools = df_service[df_service["Specification 1"].isin(expanded_codes)].copy()

            # --- Row-by-row display with dividers ---
            if not df_tools.empty:
                display_rows = []
                
                # --- Special cases with divider ---
                for sc in used_special_cases:
                    divider = pd.DataFrame({col: "" for col in df_tools.columns}, index=[0])
                    divider["Specification 1"] = f"--- {sc} ---"
                    display_rows.append(divider)
                
                    for item in special_cases[sc]:
                        item_rows = df_tools[df_tools["Specification 1"] == item]
                        if not item_rows.empty:
                            display_rows.append(item_rows)  # include all rows
                
                # --- Non-special tools ---
                non_special_mask = ~df_tools["Specification 1"].isin(sum(special_cases.values(), []))
                non_special_df = df_tools[non_special_mask]
                if not non_special_df.empty:
                    display_rows.append(non_special_df)
                
                # --- Combine ---
                if display_rows:
                    display_df = pd.concat(display_rows, ignore_index=True)
                else:
                    display_df = df_tools.copy()


                def highlight_divider(row):
                    if str(row["Specification 1"]).startswith("---"):
                        return ["background-color: red; color: white"] * len(row)
                    return [""] * len(row)

                st.subheader(f"Selected Data - Package {selected_package}, Service {selected_service}")
                st.dataframe(display_df.style.apply(highlight_divider, axis=1))

                # --- Calculation
                #Build Calculated Cost Table from display_df
                calc_df = display_df.copy()
                
                numeric_cols = [
                    "Quantity of Tools", "Total Days", "Total Months", "Total Depth (ft)",
                    "Total Survey (ft)", "Total Hours", "Discount (%)", "Daily Rate",
                    "Monthly Rate", "Depth Charge (per ft)", "Flat Charge", "Survey Charge (per ft)",
                    "Hourly Charge", "Total Flat Charge"
                ]
                
                # Ensure all numeric columns exist
                for col in numeric_cols:
                    if col not in calc_df.columns:
                        calc_df[col] = 0
                
                # --- Assign other sidebar inputs
                calc_df["Total Days"] = total_days
                calc_df["Total Months"] = total_months
                calc_df["Total Depth (ft)"] = total_depth
                calc_df["Total Survey (ft)"] = total_survey
                calc_df["Total Hours"] = total_hours
                calc_df["Discount (%)"] = discount * 100
                calc_df["Total Flat Charge"] = 0  # Start from 0
                
                # --- Define groups and their Total Flat Charge values ---
                flat_charge_groups = {
                    1: ["ECS-NMR (150DegC Max)", "PN1: PROC_NMR1", "PN2: PROC_NMR2", "PN3: PROC_NMR3",
                        "PN6: PROC_NMR6", "PE1: PROC_ES1", "PP1: PROC_PETR1", "PP6: PROC_PETR6",
                        "Dual-OBMI DSI (150DegC Max)", "PA12: PROC_ACOU14", "PI1: PROC_IMAG1", "PI2: PROC_IMAG2",
                        "PI7: PROC_IMAG7", "PI8: PROC_IMAG8", "PI9: PROC_IMAG9", "PI12: PROC_IMAG12", "PI13: PROC_IMAG13",
                        "MDT: LFA-QS-XLD-MIFA-Saturn-2MS (150DegC Max)", "PPT12: PROC_PT12",
                        "Unit, Cables & Conveyance", "DT2:RTDT_SAT"
                    ],
                    2: ["MDT: LFA-QS-XLD-MIFA-Saturn-2MS (150DegC Max)", "FP19: FPS_SPHA", "FP23: FPS_TRA"],
                    4: ["Dual-OBMI DSI (150DegC Max)", "PP7: PROC_PETR7", "PA7: PROC_ACOU6", "PA11: PROC_ACOU13",
                        "MDT: LFA-QS-XLD-MIFA-Saturn-2MS (150DegC Max)", "DT3:RTDT_PER"
                    ],
                    5: ["MDT: LFA-QS-XLD-MIFA-Saturn-2MS (150DegC Max)", "FP18: FPS_SAMP", "FP28: FPS_FCHA_1",
                        "FP33: FPS_FCHA_6", "FP34: FPS_FCHA_7", "FP11: FPS_PROB_FO", "FP26: FPS_FCON"
                    ],
                    10: ["MDT: LFA-QS-XLD-MIFA-Saturn-2MS (150DegC Max)", "FP42: FPS_PROB_XLD"],
                    50: ["XL Rock (150DegC Max)", "SC2: SC_ADD1", "SC2: SC_ADD2"]
                }
                
                # Apply Total Flat Charge per group
                for charge_value, specs in flat_charge_groups.items():
                    calc_df.loc[
                        calc_df["Specification 1"].str.upper().apply(
                            lambda x: any(spec.upper() in x for spec in specs)
                        ),
                        "Total Flat Charge"
                    ] = charge_value
                
                # --- Recalc function ---
                def recalc_costs(df):
                    df = df.copy()
                    for col in numeric_cols:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    
                    totals = []
                    for _, row in df.iterrows():
                        spec = str(row["Specification 1"])
                        if spec.startswith("---"):
                            totals.append(0)
                            continue
                        disc_fraction = row["Discount (%)"] / 100
                        total_flat = row.get("Total Flat Charge", 0)
                        operating_charge = (
                            (row["Depth Charge (per ft)"] * row["Total Depth (ft)"]) +
                            (row["Survey Charge (per ft)"] * row["Total Survey (ft)"]) +
                            (row["Flat Charge"] * total_flat) +
                            (row["Hourly Charge"] * row["Total Hours"])
                        ) * (1 - disc_fraction)
                        rental_charge = row["Quantity of Tools"] * (
                            (row["Daily Rate"] * row["Total Days"]) +
                            (row["Monthly Rate"] * row["Total Months"])
                        ) * (1 - disc_fraction)
                        totals.append(operating_charge + rental_charge)
                    df["Total (MYR)"] = totals
                    return df
                
                # Sanitize hole size for session key
                safe_hole_size = hole_size.replace('"', '_').replace('.', '_')
                calc_key = f"calc_state_{safe_hole_size}"
                
                # --- Apply Well A exceptions and sidebar quantities ---
                def apply_quantities(df, well, hole, sidebar_qty):
                    df = df.copy()
                    
                    exceptions_map = {}
                    if well == "Well A":
                        if hole == '12.25"':
                            exceptions_map = {
                                "FP18: FPS_SAMP": 11,
                                "FP19: FPS_SPHA": 4,
                                "FP23: FPS_TRA": 4,
                                "FP24: FPS_TRK": 1,
                                "FP33: FPS_FCHA_6": 1,
                                "FP34: FPS_FCHA_7": 1,
                            }
                        elif hole == '8.5"':
                            exceptions_map = {
                                "FP18: FPS_SAMP": 5,
                                "FP19: FPS_SPHA": 2,
                                "FP23: FPS_TRA": 2,
                            }
                    
                    df["clean_spec"] = df["Specification 1"].str.strip().str.upper()
                    
                    # Apply exceptions first
                    for spec_name, qty in exceptions_map.items():
                        spec_upper = spec_name.strip().upper()
                        df.loc[df["clean_spec"] == spec_upper, "Quantity of Tools"] = qty
                    
                    # Assign sidebar quantity to remaining rows
                    mask_no_exception = ~df["clean_spec"].isin([k.strip().upper() for k in exceptions_map.keys()])
                    df.loc[mask_no_exception, "Quantity of Tools"] = sidebar_qty
                    
                    df.drop(columns=["clean_spec"], inplace=True)
                    return df
                
                # --- Initial calculation ---
                calc_df = apply_quantities(calc_df, selected_well, hole_size, quantity_tools)
                calc_df = recalc_costs(calc_df)
                st.session_state[calc_key] = calc_df.copy()
                
                working_calc_df = st.session_state[calc_key].copy()
                
                # Include any new rows
                missing_specs = set(calc_df["Specification 1"]) - set(working_calc_df["Specification 1"])
                if missing_specs:
                    working_calc_df = pd.concat(
                        [working_calc_df, calc_df[calc_df["Specification 1"].isin(missing_specs)]],
                        ignore_index=True
                    )
                
                # Apply sidebar inputs and recalc
                working_calc_df["Total Days"] = total_days
                working_calc_df["Total Months"] = total_months
                working_calc_df["Total Depth (ft)"] = total_depth
                working_calc_df["Total Survey (ft)"] = total_survey
                working_calc_df["Total Hours"] = total_hours
                working_calc_df["Discount (%)"] = discount * 100
                
                working_calc_df = apply_quantities(working_calc_df, selected_well, hole_size, quantity_tools)
                working_calc_df = recalc_costs(working_calc_df).reset_index(drop=True)
                
                # --- Display editable table ---
                edited_df = st.data_editor(
                    working_calc_df,
                    num_rows="dynamic",
                    key=f"calc_editor_{safe_hole_size}",
                )
                
                # --- Reapply exceptions after user edits ---
                edited_df = apply_quantities(edited_df, selected_well, hole_size, quantity_tools)
                updated_calc_df = recalc_costs(edited_df)
                st.session_state[calc_key] = updated_calc_df
                
                # --- Section total ---
                section_total = updated_calc_df["Total (MYR)"].sum()
                section_totals[hole_size] = section_total
                st.write(f"### 💵 Section Total for {hole_size}\" Hole: {section_total:,.2f}")

                # --- Identify special tools automatically ---
                special_cases_section = {}
                for charge_value in updated_calc_df["Total Flat Charge"].unique():
                    if charge_value > 0:
                        tools = updated_calc_df.loc[updated_calc_df["Total Flat Charge"] == charge_value, "Specification 1"].tolist()
                        section_name = f"FlatCharge_{charge_value}"
                        special_cases_section[section_name] = tools
                used_special_cases = list(special_cases_section.keys())

                
                # Store for Excel download
                all_calc_dfs_for_excel.append((hole_size, used_special_cases, updated_calc_df, special_cases_section))




    # --- Grand Total ---
    if section_totals:
        grand_total = sum(section_totals.values())
        st.success(f"🏆 Grand Total Price (MYR): {grand_total:,.2f}")

# --- Excel Download ---
# --- Excel Download ---
if st.button("Download Cost Estimate Excel"):
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for hole_size, used_special_cases, df_tools_section, special_cases_section in all_calc_dfs_for_excel:
            sheet_name = f'{hole_size}" Hole'
            wb = writer.book
            ws = wb.create_sheet(title=sheet_name)

            # --- Header Rows ---
            ws.merge_cells("B2:B4"); ws["B2"]="Reference"
            ws.merge_cells("C2:C4"); ws["C2"]="Specification 1"
            ws.merge_cells("D2:D4"); ws["D2"]="Specification 2"
            ws.merge_cells("E2:J2"); ws["E2"]="Unit Price"
            ws.merge_cells("E3:F3"); ws["E3"]="Rental Price"
            ws.merge_cells("G3:J3"); ws["G3"]="Operating Charge"
            ws["E4"]="Daily Rate"; ws["F4"]="Monthly Rate"; ws["G4"]="Depth Charge (per ft)"
            ws["H4"]="Survey Charge (per ft)"; ws["I4"]="Flat Charge"; ws["J4"]="Hourly Charge"

            ws.merge_cells("K2:Q2"); ws["K2"] = "Operation Estimated"
            ws.merge_cells("K3:K4"); ws["K3"] = "Quantity of Tools"
            ws.merge_cells("L3:M3"); ws["L3"] = "Rental Parameters"
            ws["L4"] = "Total Days"; ws["M4"] = "Total Months"
            ws.merge_cells("N3:Q3"); ws["N3"] = "Operating Parameters"
            ws["N4"] = "Total Depth (ft)"; ws["O4"] = "Total Survey (ft)"
            ws["P4"] = "Total Flat Charge (ft)"; ws["Q4"] = "Total Hours"
            ws.merge_cells("R2:R4"); ws["R2"] = "Discount (%)"
            ws.merge_cells("S2:S4"); ws["S2"] = "Total (MYR)"
            ws.merge_cells("T2:T4"); ws["T2"] = "Grand Total Price (MYR)"
            ws.merge_cells("U2:U3"); ws["U2"] = "Break Down"; ws["U4"] = "Rental Charge (MYR)"
            ws.merge_cells("V2:V3"); ws["V2"] = "Break Down"; ws["V4"] = "Operating Charge (MYR)"

            # --- Apply Colors ---
            white_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            light_green_fill = PatternFill(start_color="CCCC99", end_color="CCCC99", fill_type="solid")
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # White headers
            for cell in ["B2","B3","B4","C2","C3","C4","D2","D3","D4","R2","R3","R4","S2","S3","S4","T2","T3","T4","U2","U3","V2","V3"]:
                ws[cell].fill = white_fill
                ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Light Green headers
            for cell in ["E2","E3","E4","F2","F3","F4","G2","G3","G4","H4","I4","J4","U4","V4"]:
                ws[cell].fill = light_green_fill
                ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Blue headers
            for cell in ["K2","K3","K4","L3","L4","M3","M4","N3","N4","O4","P4","Q4"]:
                ws[cell].fill = blue_fill
                ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # --- Insert Data ---
            current_row = 5
            first_data_row = current_row  # Track first row for Grand Total formula

            # --- Insert Special Tools by Section ---
            for section_name in used_special_cases:
                if section_name not in special_cases_map["STANDARD WELLS"]:
                    continue
                items_in_section = special_cases_map["STANDARD WELLS"][section_name]

                # Section header
                ws[f"B{current_row}"] = f'{hole_size}in Section: {section_name}'
                ws[f"B{current_row}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws[f"B{current_row}"].alignment = Alignment(horizontal="center")
                current_row += 1

                # Items under this section
                for item in items_in_section:
                    item_rows = df_tools_section[df_tools_section["Specification 1"] == item]
                    if not item_rows.empty:
                        item_row = item_rows.iloc[0]
                        ws[f"B{current_row}"] = item_row.get("Reference","")
                        ws[f"C{current_row}"] = item_row.get("Specification 1","")
                        ws[f"D{current_row}"] = item_row.get("Specification 2","")
                        ws[f"E{current_row}"] = item_row.get("Daily Rate",0)
                        ws[f"F{current_row}"] = item_row.get("Monthly Rate",0)
                        ws[f"G{current_row}"] = item_row.get("Depth Charge (per ft)",0)
                        ws[f"H{current_row}"] = item_row.get("Survey Charge (per ft)",0)
                        ws[f"I{current_row}"] = item_row.get("Flat Charge",0)
                        ws[f"J{current_row}"] = item_row.get("Hourly Charge",0)

                        # Operation Estimated values
                        qty = st.session_state.get(f"qty_{hole_size}",0)
                        total_days = st.session_state.get(f"days_{hole_size}",0)
                        total_months = st.session_state.get(f"months_{hole_size}",0)
                        total_depth = st.session_state.get(f"depth_{hole_size}",0)
                        total_survey = st.session_state.get(f"survey_{hole_size}",0)
                        total_hours = st.session_state.get(f"hours_{hole_size}",0)
                        discount_pct = st.session_state.get(f"disc_{hole_size}",0)

                        ws[f"K{current_row}"] = qty
                        ws[f"L{current_row}"] = total_days
                        ws[f"M{current_row}"] = total_months
                        ws[f"N{current_row}"] = total_depth
                        ws[f"O{current_row}"] = total_survey
                        ws[f"P{current_row}"] = ws[f"I{current_row}"].value if ws[f"I{current_row}"].value else 0
                        ws[f"Q{current_row}"] = total_hours
                        ws[f"R{current_row}"] = discount_pct * 100

                        rental_charge = qty * ((item_row.get("Daily Rate",0)*total_days) + (item_row.get("Monthly Rate",0)*total_months))*(1-discount_pct)
                        operating_charge = ((item_row.get("Depth Charge (per ft)",0)*total_depth)+
                                            (item_row.get("Survey Charge (per ft)",0)*total_survey)+
                                            (item_row.get("Flat Charge",0))+ 
                                            (item_row.get("Hourly Charge",0)*total_hours))*(1-discount_pct)
                        total_myr = rental_charge + operating_charge

                        ws[f"S{current_row}"] = total_myr
                        ws[f"U{current_row}"] = rental_charge
                        ws[f"V{current_row}"] = operating_charge

                        current_row += 1

            # --- Insert Non-Special Tools ---
            for item in df_tools_section["Specification 1"]:
                if item not in sum(special_cases_section.values(), []):
                    item_rows = df_tools_section[df_tools_section["Specification 1"] == item]
                    if not item_rows.empty:
                        item_row = item_rows.iloc[0]
                        ws[f"B{current_row}"] = item_row.get("Reference","")
                        ws[f"C{current_row}"] = item_row.get("Specification 1","")
                        ws[f"D{current_row}"] = item_row.get("Specification 2","")
                        ws[f"E{current_row}"] = item_row.get("Daily Rate",0)
                        ws[f"F{current_row}"] = item_row.get("Monthly Rate",0)
                        ws[f"G{current_row}"] = item_row.get("Depth Charge (per ft)",0)
                        ws[f"H{current_row}"] = item_row.get("Survey Charge (per ft)",0)
                        ws[f"I{current_row}"] = item_row.get("Flat Charge",0)
                        ws[f"J{current_row}"] = item_row.get("Hourly Charge",0)

                        qty = st.session_state.get(f"qty_{hole_size}",0)
                        total_days = st.session_state.get(f"days_{hole_size}",0)
                        total_months = st.session_state.get(f"months_{hole_size}",0)
                        total_depth = st.session_state.get(f"depth_{hole_size}",0)
                        total_survey = st.session_state.get(f"survey_{hole_size}",0)
                        total_hours = st.session_state.get(f"hours_{hole_size}",0)
                        discount_pct = st.session_state.get(f"disc_{hole_size}",0)

                        ws[f"K{current_row}"] = qty
                        ws[f"L{current_row}"] = total_days
                        ws[f"M{current_row}"] = total_months
                        ws[f"N{current_row}"] = total_depth
                        ws[f"O{current_row}"] = total_survey
                        ws[f"P{current_row}"] = ws[f"I{current_row}"].value if ws[f"I{current_row}"].value else 0
                        ws[f"Q{current_row}"] = total_hours
                        ws[f"R{current_row}"] = discount_pct * 100

                        rental_charge = qty * ((item_row.get("Daily Rate",0)*total_days) + (item_row.get("Monthly Rate",0)*total_months))*(1-discount_pct)
                        operating_charge = ((item_row.get("Depth Charge (per ft)",0)*total_depth)+
                                            (item_row.get("Survey Charge (per ft)",0)*total_survey)+
                                            (item_row.get("Flat Charge",0))+ 
                                            (item_row.get("Hourly Charge",0)*total_hours))*(1-discount_pct)
                        total_myr = rental_charge + operating_charge

                        ws[f"S{current_row}"] = total_myr
                        ws[f"U{current_row}"] = rental_charge
                        ws[f"V{current_row}"] = operating_charge

                        current_row += 1

            # --- Grand Total ---
            ws[f"T{first_data_row}"] = f"=SUM(S{first_data_row}:S{current_row-1})"
            ws[f"T{first_data_row}"].alignment = Alignment(horizontal="center")

    output.seek(0)
    st.download_button(
        "Download Cost Estimate Excel",
        data=output,
        file_name="Cost_Estimate.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )







































